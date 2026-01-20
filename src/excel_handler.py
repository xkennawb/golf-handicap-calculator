"""
Excel Handler for Handicap Tracking
Reads and writes player handicap data to Excel spreadsheet
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os


class ExcelHandler:
    def __init__(self, file_path='handicaps.xlsx'):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        
    def initialize_workbook(self):
        """Create a new workbook with proper headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Handicaps"
        
        # Headers
        headers = ['Date', 'Player', 'Course', 'Gross Score', 'Stableford', 'Par', 
                   'Score to Par', 'Playing HC', 'Net Score', 'Differential', 
                   'Weather Factor', 'Weather Conditions', 'Current Index', 'Notes']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 18  # Player
        ws.column_dimensions['C'].width = 25  # Course
        ws.column_dimensions['D'].width = 12  # Gross Score
        ws.column_dimensions['E'].width = 12  # Stableford
        ws.column_dimensions['F'].width = 8   # Par
        ws.column_dimensions['G'].width = 12  # Score to Par
        ws.column_dimensions['H'].width = 12  # Playing HC
        ws.column_dimensions['I'].width = 12  # Net Score
        ws.column_dimensions['J'].width = 12  # Differential
        ws.column_dimensions['K'].width = 14  # Weather Factor
        ws.column_dimensions['L'].width = 25  # Weather Conditions
        ws.column_dimensions['M'].width = 14  # Current Index
        ws.column_dimensions['N'].width = 30  # Notes
        
        wb.save(self.file_path)
        return wb
    
    def load_or_create_workbook(self):
        """Load existing workbook or create new one"""
        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
            self.sheet = self.workbook.active
        else:
            self.workbook = self.initialize_workbook()
            self.sheet = self.workbook.active
        
        return self.workbook
    
    def add_round(self, round_date, course_name, player_results, weather_data):
        """
        Add a round's results to the spreadsheet
        player_results: list of dicts from RoundAnalyzer
        weather_data: dict with weather conditions
        """
        if not self.workbook:
            self.load_or_create_workbook()
        
        # Find the next empty row
        next_row = self.sheet.max_row + 1
        
        # Format weather description
        weather_desc = f"{weather_data['description']}, {weather_data['temperature']:.1f}°C, Wind {weather_data['wind_speed']:.0f}km/h"
        
        # Add each player's results
        for result in player_results:
            row_data = [
                round_date.strftime('%Y-%m-%d'),
                result['name'],
                course_name,
                result['gross_score'],
                result['stableford_points'],
                result['par'],
                result['score_to_par'],
                result['playing_handicap'],
                result['net_score'],
                result['score_differential'],
                result['weather_factor'],
                weather_desc,
                result['current_handicap_index'],
                f"9-hole round"
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = self.sheet.cell(row=next_row, column=col, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            next_row += 1
        
        # Save the workbook
        self.workbook.save(self.file_path)
        
        return next_row - self.sheet.max_row - 1  # Number of rows added
    
    def get_player_history(self, player_name, num_rounds=20):
        """
        Get recent score history for a player
        Returns list of differentials for handicap calculation
        """
        if not self.workbook:
            self.load_or_create_workbook()
        
        differentials = []
        weather_factors = []
        
        # Iterate through rows (skip header)
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == player_name:  # Player name column
                if row[8] is not None:  # Differential column
                    differentials.append(float(row[8]))
                    if row[9] is not None:  # Weather factor column
                        weather_factors.append(float(row[9]))
                    else:
                        weather_factors.append(1.0)
        
        # Return most recent rounds
        return differentials[-num_rounds:], weather_factors[-num_rounds:]
    
    def get_latest_handicaps(self):
        """
        Get the most recent handicap index for each player
        Returns dict: {player_name: handicap_index}
        """
        if not self.workbook:
            self.load_or_create_workbook()
        
        latest_handicaps = {}
        
        # Iterate through rows in reverse to get latest first
        for row in reversed(list(self.sheet.iter_rows(min_row=2, values_only=True))):
            player_name = row[1]
            handicap_index = row[11]  # Current Index column
            
            if player_name and handicap_index is not None:
                if player_name not in latest_handicaps:
                    latest_handicaps[player_name] = float(handicap_index)
        
        return latest_handicaps
    
    def get_year_stats(self, year=None):
        """
        Get comprehensive statistics for all players for the specified year
        Returns dict with player stats
        """
        if not self.workbook:
            self.load_or_create_workbook()
        
        if year is None:
            year = datetime.now().year
        
        player_stats = {}
        
        # Iterate through all rows
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            date_str = row[0]
            if not date_str:
                continue
            
            # Parse date
            if isinstance(date_str, str):
                try:
                    round_date = datetime.strptime(date_str, '%Y-%m-%d')
                except:
                    continue
            else:
                round_date = date_str
            
            # Filter by year
            if round_date.year != year:
                continue
            
            player_name = row[1]
            gross_score = row[3]
            stableford = row[4]
            net_score = row[8]
            current_index = row[12]  # Current Index column
            
            if not player_name or gross_score is None:
                continue
            
            # Initialize player stats
            if player_name not in player_stats:
                player_stats[player_name] = {
                    'games_played': 0,
                    'games_won': 0,
                    'gross_scores': [],
                    'stableford_scores': [],
                    'net_scores': [],
                    'best_gross': float('inf'),
                    'best_stableford': 0,
                    'current_index': None
                }
            
            stats = player_stats[player_name]
            stats['games_played'] += 1
            stats['gross_scores'].append(gross_score)
            stats['stableford_scores'].append(stableford if stableford else 0)
            stats['net_scores'].append(net_score if net_score else gross_score)
            
            # Store the most recent handicap index for this player
            if current_index is not None:
                stats['current_index'] = current_index
            
            if gross_score < stats['best_gross']:
                stats['best_gross'] = gross_score
            
            if stableford and stableford > stats['best_stableford']:
                stats['best_stableford'] = stableford
        
        # Determine winners for each round (lowest net score)
        round_winners = self._get_round_winners(year)
        for winner in round_winners:
            if winner in player_stats:
                player_stats[winner]['games_won'] += 1
        
        # Calculate averages and win percentages
        for player_name, stats in player_stats.items():
            if stats['gross_scores']:
                stats['avg_gross'] = round(sum(stats['gross_scores']) / len(stats['gross_scores']), 1)
            else:
                stats['avg_gross'] = 0
            
            if stats['stableford_scores']:
                stats['avg_stableford'] = round(sum(stats['stableford_scores']) / len(stats['stableford_scores']), 1)
            else:
                stats['avg_stableford'] = 0
            
            if stats['games_played'] > 0:
                stats['win_percentage'] = round((stats['games_won'] / stats['games_played']) * 100, 1)
            else:
                stats['win_percentage'] = 0
        
        return player_stats
    
    def _get_round_winners(self, year):
        """
        Determine the winner of each round (highest Stableford score)
        Returns list of winner names
        """
        if not self.workbook:
            self.load_or_create_workbook()
        
        rounds = {}  # date -> [(player, net_score)]
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            date_str = row[0]
            if not date_str:
                continue
            
            # Parse date
            if isinstance(date_str, str):
                try:
                    round_date = datetime.strptime(date_str, '%Y-%m-%d')
                except:
                    continue
            else:
                round_date = date_str
            
            if round_date.year != year:
                continue
            
            player_name = row[1]
            stableford = row[4]  # Stableford column
            
            if not player_name or stableford is None:
                continue
            
            date_key = round_date.strftime('%Y-%m-%d')
            if date_key not in rounds:
                rounds[date_key] = []
            
            rounds[date_key].append((player_name, stableford))
        
        # Find winner of each round (highest Stableford)
        winners = []
        for date_key, players in rounds.items():
            if players:
                winner = max(players, key=lambda x: x[1])
                winners.append(winner[0])
        
        return winners
    
    def get_summary_for_whatsapp(self, player_results, course_name, round_date, include_year_stats=True):
        """
        Generate a formatted summary message for WhatsApp
        """
        message = f"Golf Handicaps Update\n"
        message += f"{round_date.strftime('%A, %B %d, %Y')}\n"
        message += f"{course_name}\n\n"
        
        # Sort by Stableford score (highest wins)
        sorted_results = sorted(player_results, key=lambda x: x['stableford_points'], reverse=True)
        
        message += "Today's Results:\n"
        for i, result in enumerate(sorted_results, start=1):
            medal = "[1st]" if i == 1 else "[2nd]" if i == 2 else "[3rd]" if i == 3 else "     "
            message += f"{medal} {result['name']}\n"
            message += f"   Gross: {result['gross_score']} ({result['score_to_par']:+d})"
            message += f" | Stableford: {result['stableford_points']}\n"
            message += f"   Net: {result['net_score']} ({result['net_to_par']:+d})"
            message += f" | Index: {result['current_handicap_index']}\n"
        
        message += f"\nWeather Factor: {player_results[0]['weather_factor']}x\n"
        
        # Add year-to-date stats
        if include_year_stats:
            year_stats = self.get_year_stats(round_date.year)
            
            if year_stats:
                message += f"\n{round_date.year} Season Stats:\n"
                
                # Sort by win percentage
                sorted_stats = sorted(
                    year_stats.items(),
                    key=lambda x: (x[1]['win_percentage'], x[1]['avg_stableford']),
                    reverse=True
                )
                
                for player_name, stats in sorted_stats:
                    message += f"\n{player_name}"
                    if stats.get('current_index') is not None:
                        message += f" (Index: {stats['current_index']})"
                    message += ":\n"
                    message += f"  Games: {stats['games_played']} | Wins: {stats['games_won']} ({stats['win_percentage']}%)\n"
                    message += f"  Avg Gross: {stats['avg_gross']:.1f} | Avg Stableford: {stats['avg_stableford']:.1f}\n"
                    message += f"  Best Gross: {stats['best_gross']}"
                    if stats['best_stableford'] > 0:
                        message += f" | Best Stableford: {stats['best_stableford']}"
                    message += "\n"
        
        return message


if __name__ == "__main__":
    # Test the Excel handler
    handler = ExcelHandler('test_handicaps.xlsx')
    handler.load_or_create_workbook()
    
    # Sample data
    sample_results = [
        {
            'name': 'Bruce Kennaway',
            'gross_score': 47,
            'stableford_points': 13,
            'par': 36,
            'score_to_par': 11,
            'playing_handicap': 8,
            'net_score': 39,
            'net_to_par': 3,
            'score_differential': 14.5,
            'current_handicap_index': 19.0,
            'weather_factor': 1.08
        }
    ]
    
    sample_weather = {
        'description': 'partly cloudy',
        'temperature': 22.5,
        'wind_speed': 18
    }
    
    round_date = datetime(2025, 12, 5)
    course_name = "Warringah Golf Club"
    
    handler.add_round(round_date, course_name, sample_results, sample_weather)
    
    print("✅ Sample data added to spreadsheet")
    print(f"📊 File saved: {handler.file_path}")
    
    # Test WhatsApp summary
    message = handler.get_summary_for_whatsapp(sample_results, course_name, round_date)
    print("\n" + "="*50)
    print("WhatsApp Message Preview:")
    print("="*50)
    print(message)
